import openpyxl
import requests

def standardize_address(address):
    api_key = 'API KEY'
    base_url = 'https://dev.virtualearth.net/REST/v1/Locations'

    query_params = {
        'query': address,
        'includeNeighborhood': 1,
        'maxResults': 1,
        'key': api_key,
        'userMapView': '40.6829,-74.0479,40.8790,-73.9067'
    }

    response = requests.get(base_url, params=query_params)
    
    # Check if response is successful
    if response.status_code == 200:
        try:
            data = response.json()
        except ValueError:
            print("Error: Unable to decode response JSON.")
            return None
    else:
        print(f"Error: Request failed with status code {response.status_code}")
        return None

    if 'resourceSets' in data and len(data['resourceSets']) > 0 and 'resources' in data['resourceSets'][0]:
        resources = data['resourceSets'][0]['resources']
        if resources:
            location = resources[0]
            standardized_address = [
                location.get('address',{}).get('addressLine',''),
                location.get('address',{}).get('locality',''),
                location.get('address',{}).get('adminDistrict',''),
                location.get('address',{}).get('countryRegion',''),
                location.get('address',{}).get('postalCode',''),
            ]
            return standardized_address
    return None

def get_aka_names(address):
    local_ip = '192.168.x.x'  # Replace with your local IP address
    port = '8080'  # Replace with the port number on which Nominatim is running
    bound_box = "-74.0479,40.6829,-73.9067,40.8790"
    
    # Update the URL to point to the local Nominatim service
    url = f"http://{local_ip}:{port}/search?format=json&q={address}&bounded=1&viewbox={bound_box}"
    
    response = requests.get(url)
    
    # Check if response is successful
    if response.status_code == 200:
        try:
            data = response.json()
        except ValueError:
            print("Error: Unable to decode response JSON.")
            return set()
    else:
        print(f"Error: Request failed with status code {response.status_code}")
        return set()

    aka_names = set()
    for entry in data:
        if 'display_name' in entry:
            aka_names.add(entry['display_name'])
    return aka_names

def process_excel(input_file, output_file):
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active

    output_wb = openpyxl.Workbook()
    output_ws = output_wb.active
    output_ws.append(["Original Address", 'Street Address','Locality','Admin District','Country Region','Postal Code', 'AKA Count', 'AKA Names'])

    for row in ws.iter_rows(min_row=2, values_only=True):
        original_address = row[0]
        standardized_address = standardize_address(original_address)
        if standardized_address:
            aka_names = get_aka_names(standardized_address)
            aka_count = len(aka_names)
            output_ws.append([original_address] + standardized_address + [aka_count, " | ".join(aka_names)])
        else:
            output_ws.append([original_address] + [''] * 5 + [0, ''])

    output_wb.save(output_file)

if __name__ == "__main__":
    input_file = "manhattan test.xlsx"
    output_file = "manhattan_test_output.xlsx"
    process_excel(input_file, output_file)
