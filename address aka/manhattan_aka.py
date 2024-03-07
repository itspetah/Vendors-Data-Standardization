import requests
import openpyxl

# Define API endpoint and user agent
nominatim_url = "https://nominatim.openstreetmap.org/search"
user_agent = "your_application_name/version (your_email@example.com)"  # Replace with your details

# Define output file name
output_file = "manhattan_addresses.xlsx"

# Function to call Nominatim API and retrieve address details
def get_address_details(address):
    params = {
        "q": address,
        "format": "json",
        "user-agent": user_agent
    }
    response = requests.get(nominatim_url, params=params)
    return response.json()

# Function to process address details and extract relevant information
def process_address(address_data):
    if address_data:
        address_components = address_data[0]["address"]
        street = address_components.get("road", "")
        housenumber = address_components.get("housenumber", "")
        postcode = address_components.get("postcode", "")
        if street and housenumber:
            full_address = f"{housenumber} {street}, Manhattan, NYC, {postcode}"
            aka_names = []
            for name in address_data:
                if name["display_name"] != full_address:
                    aka_names.append(name["display_name"])
            return full_address, len(aka_names), aka_names
    return None, None, None

# Create a new Excel workbook
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.cell(row=1, column=1).value = "Full Address"
sheet.cell(row=1, column=2).value = "AKA Name Count"
sheet.cell(row=1, column=3).value = "AKA Names"

# Define path to the file containing addresses (replace with your actual path)
address_file = "path/to/your/addresses.txt"

# Open the address file and iterate through each line
with open(address_file, "r") as f:
    for address in f:
        # Remove trailing newline character from the address
        address = address.strip()
        full_address, aka_name_count, aka_names = process_address(get_address_details(address))
        if full_address:
            row = sheet.max_row + 1
            sheet.cell(row=row, column=1).value = full_address
            sheet.cell(row=row, column=2).value = aka_name_count
            sheet.cell(row=row, column=3).value = ", ".join(aka_names)

# Save the Excel workbook
workbook.save(output_file)

print(f"Addresses and AKA names saved to '{output_file}'.")
