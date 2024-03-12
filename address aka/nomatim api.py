import requests
import openpyxl
import time

# Define API endpoint and user agent
nominatim_url = "https://nominatim.openstreetmap.org/search"
user_agent = "your_application_name/version (your_email@example.com)"  # Replace with your details

output_file = "manhattan_addresses.xlsx"

def get_address_details(address):
    params = {
        "q": address,
        "format": "json",
        "user-agent": user_agent
    }
    response = requests.get(nominatim_url, params=params)
    return response.json()

def process_address(address_data):
    if address_data:
        try:
            address_components = address_data[0]["address"]
            street = address_components.get("road", "")
            housenumber = address_components.get("housenumber", "")
            postcode = address_components.get("postcode", "")
            if street and housenumber:
                full_address = f"{housenumber} {street}, Manhattan, NYC, {postcode}"
                aka_names = []
                for name in address_data:
                    if name.get("display_name") != full_address:
                        aka_names.append(name.get("display_name", ""))
                return full_address, len(aka_names), aka_names
        except KeyError as e:
            print(f"KeyError: {e}. Address data might be missing or in unexpected format.")
    # If address processing fails, return the input address as full address
    return address, None, None

# Create a new Excel workbook
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.cell(row=1, column=1).value = "Input Address"
sheet.cell(row=1, column=2).value = "Full Address"
sheet.cell(row=1, column=3).value = "AKA Name Count"
sheet.cell(row=1, column=4).value = "AKA Names"

address_file = "Manhattan Addresses.xlsx"

with open(address_file, "r", encoding="utf-8") as f:
    for row_index, address in enumerate(f, start=2):  # Start from row 2
        address = address.strip()
        full_address, aka_name_count, aka_names = process_address(get_address_details(address))
        if full_address:
            sheet.cell(row=row_index, column=1).value = address.encode('utf-8', 'ignore').decode()  # Encode the address string
            sheet.cell(row=row_index, column=2).value = full_address.encode('utf-8', 'ignore').decode()  # Encode the address string
            sheet.cell(row=row_index, column=3).value = aka_name_count
            sheet.cell(row=row_index, column=4).value = ", ".join(aka_names)
        else:
            # If address processing fails, add input address to the output file
            sheet.cell(row=row_index, column=1).value = address.encode('utf-8', 'ignore').decode()  # Encode the address string
            print(f"Unable to process address: {address}")
        time.sleep(1)  # Add delay to avoid rate limiting

workbook.save(output_file)

print(f"Addresses and AKA names saved to '{output_file}'.")
